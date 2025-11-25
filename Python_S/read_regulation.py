import pandas as pd
import os
from openpyxl.utils import get_column_letter, column_index_from_string

def extract_data_from_range():
    """
    从Country Overview工作表的E7到GZ95范围提取数据，去重后返回一个数组
    
    Returns:
        list: 去重后的数据数组
    """
    try:
        # 文件路径
        file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'DataStorage', 'Regulation.xlxs.xlsx')
        
        print(f"正在读取文件: {file_path}")
        print("正在提取E7到GZ95范围的数据...")
        
        # 计算列索引
        start_col = column_index_from_string('E')  # E列对应索引4（0开始）
        end_col = column_index_from_string('GZ')   # GZ列对应索引289（0开始）
        start_row = 6  # 第7行对应索引6（0开始）
        end_row = 94   # 第95行对应索引94（0开始）
        
        # 读取整个工作表
        df = pd.read_excel(file_path, sheet_name='Country Overview', engine='openpyxl')
        
        # 确保数据范围有效
        if df.shape[0] <= end_row or df.shape[1] <= end_col:
            print(f"警告：工作表实际尺寸 ({df.shape[0]}行 × {df.shape[1]}列) 小于请求的范围 (E7:GZ95)")
            # 调整范围以适应实际数据
            end_row = min(end_row, df.shape[0] - 1)
            end_col = min(end_col, df.shape[1] - 1)
        
        # 提取指定范围的数据
        range_df = df.iloc[start_row:end_row+1, start_col:end_col+1]
        
        print(f"已提取范围: {get_column_letter(start_col+1)}{start_row+1}到{get_column_letter(end_col+1)}{end_row+1}")
        print(f"提取的数据维度: {range_df.shape[0]}行 × {range_df.shape[1]}列")
        
        # 将所有非空值提取到一个集合中进行去重
        unique_values = set()
        
        # 遍历所有单元格
        for _, row in range_df.iterrows():
            for cell_value in row:
                # 检查值是否非空且不是NaN
                if pd.notna(cell_value) and cell_value != '':
                    # 转换为字符串并去除首尾空格
                    str_value = str(cell_value).strip()
                    if str_value:
                        unique_values.add(str_value)
        
        # 转换为排序后的列表
        result_array = sorted(list(unique_values))
        
        print(f"去重后的数据总数: {len(result_array)}")
        print("\n前20个数据样本:")
        for i, value in enumerate(result_array[:20], 1):
            print(f"{i}. {value}")
        
        if len(result_array) > 20:
            print(f"... 以及其他 {len(result_array) - 20} 个数据")
        
        return result_array
    
    except Exception as e:
        print(f"提取数据时出错: {e}")
        return []

def read_country_overview():
    """
    读取Regulation.xlxs.xlsx文件的Country Overview工作表内容
    分析ADAS系统在不同国家的法规要求
    """
    try:
        # 文件路径
        file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'DataStorage', 'Regulation.xlxs.xlsx')
        
        print(f"正在读取文件: {file_path}")
        
        # 读取Excel文件的Country Overview工作表
        df = pd.read_excel(file_path, sheet_name='Country Overview')
        
        # 显示工作表的基本信息
        print("\n=== 工作表基本信息 ===")
        print(f"工作表名称: Country Overview")
        print(f"数据维度: {df.shape[0]}行 × {df.shape[1]}列")
        
        print("\n=== 列名信息 ===")
        for i, col in enumerate(df.columns, 1):
            print(f"{i}. {col}")
        
        # 显示前5行数据作为样例
        print("\n=== 数据样例（前5行）===")
        print(df.head().to_string())
        
        # 分析国家/地区信息
        country_col = None
        for col in ['Country', '国家', 'Country/Region', '国家/地区']:
            if col in df.columns:
                country_col = col
                break
        
        if country_col:
            print(f"\n=== {country_col} 分析 ===")
            countries = df[country_col].dropna().unique()
            print(f"包含的国家/地区数量: {len(countries)}")
            print(f"国家/地区列表: {', '.join(map(str, countries))}")
        
        # 分析ADAS功能相关列
        print("\n=== ADAS功能相关分析 ===")
        function_cols = []
        for col in df.columns:
            if any(keyword in str(col).lower() for keyword in ['adas', 'function', 'feature', '功能']):
                function_cols.append(col)
        
        if function_cols:
            print(f"发现 {len(function_cols)} 个可能与ADAS功能相关的列:")
            for col in function_cols:
                unique_vals = df[col].dropna().unique()
                print(f"- {col}: {len(unique_vals)} 个唯一值")
                if len(unique_vals) <= 10:  # 如果唯一值较少，全部显示
                    print(f"  唯一值: {', '.join(map(str, unique_vals))}")
        else:
            print("未发现明确的ADAS功能相关列")
        
        # 分析法规相关列
        print("\n=== 法规标准相关分析 ===")
        regulation_cols = []
        for col in df.columns:
            if any(keyword in str(col).lower() for keyword in ['regulation', '法规', 'standard', '标准', 'requirement', '要求']):
                regulation_cols.append(col)
        
        if regulation_cols:
            print(f"发现 {len(regulation_cols)} 个可能与法规标准相关的列:")
            for col in regulation_cols:
                unique_vals = df[col].dropna().unique()
                print(f"- {col}: {len(unique_vals)} 个唯一值")
                if len(unique_vals) <= 10:  # 如果唯一值较少，全部显示
                    print(f"  唯一值: {', '.join(map(str, unique_vals))}")
        else:
            print("未发现明确的法规标准相关列")
        
        # 分析法规实施状态
        print("\n=== 法规实施状态分析 ===")
        status_cols = []
        for col in df.columns:
            if any(keyword in str(col).lower() for keyword in ['status', '状态', 'effective', '生效', 'implementation', '实施']):
                status_cols.append(col)
        
        if status_cols:
            for col in status_cols:
                print(f"\n{col} 分布:")
                value_counts = df[col].value_counts()
                for val, count in value_counts.items():
                    print(f"- {val}: {count} 次出现")
        
        # 提供数据关系分析建议
        print("\n=== 数据关系分析建议 ===")
        if country_col and function_cols and regulation_cols:
            print("1. 可以分析每个国家/地区对不同ADAS功能的法规要求")
            print("2. 可以统计各ADAS功能在不同国家/地区的法规合规状态")
            print("3. 可以比较不同地区间法规要求的差异")
        
        return df
    except Exception as e:
        print(f"读取文件时出错: {e}")
        # 尝试处理可能的编码问题或其他常见错误
        try:
            # 尝试不同的读取方式
            df = pd.read_excel(file_path, sheet_name='Country Overview', engine='openpyxl')
            print("使用openpyxl引擎成功读取文件")
            return df
        except Exception as e2:
            print(f"使用openpyxl引擎也失败: {e2}")
        return None

if __name__ == "__main__":
    print("=== ADAS法规数据提取工具 ===")
    print("选项1: 提取E7到GZ95范围的去重数据")
    print("选项2: 执行完整的工作表分析")
    
    choice = input("请选择操作 (1/2): ")
    
    if choice == '1':
        print("\n正在执行数据范围提取...")
        result = extract_data_from_range()
        print(f"\n数据提取完成！共获取 {len(result)} 个不重复的数据项。")
        
        # 询问是否保存结果到文件
        save_choice = input("是否保存结果到文件？(y/n): ")
        if save_choice.lower() == 'y':
            try:
                output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'DataStorage', 'extracted_unique_data.txt')
                with open(output_path, 'w', encoding='utf-8') as f:
                    for item in result:
                        f.write(f"{item}\n")
                print(f"结果已保存到: {output_path}")
            except Exception as e:
                print(f"保存文件时出错: {e}")
    
    elif choice == '2':
        print("\n正在执行完整的工作表分析...")
        df = read_country_overview()
        if df is not None:
            print("\n分析完成！您可以根据输出结果进一步了解ADAS系统在不同国家的法规要求。")
    
    else:
        print("无效的选择，程序退出。")
        exit(1)