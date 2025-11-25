import os
import pandas as pd
from openpyxl import load_workbook

def write_unique_data_to_excel():
    """
    读取extracted_unique_data.txt中的不重复内容，并将其写入到database.xlsx文件的REGULATION工作表中，从C1单元格开始。
    """
    # 设置文件路径
    data_storage_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'DataStorage')
    text_file_path = os.path.join(data_storage_path, 'extracted_unique_data.txt')
    excel_file_path = os.path.join(data_storage_path, 'database.xlsx')
    
    try:
        # 读取文本文件中的内容
        with open(text_file_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 去除每行的换行符和前后空白
        unique_data = [line.strip() for line in lines if line.strip()]
        
        print(f"成功读取到 {len(unique_data)} 条不重复数据")
        
        # 检查Excel文件是否存在
        if not os.path.exists(excel_file_path):
            print(f"错误：Excel文件 {excel_file_path} 不存在")
            return
        
        # 使用openpyxl打开Excel文件
        book = load_workbook(excel_file_path)
        
        # 检查REGULATION工作表是否存在
        if 'REGULATION' not in book.sheetnames:
            print("错误：REGULATION工作表不存在")
            book.close()
            return
        
        # 获取REGULATION工作表
        worksheet = book['REGULATION']
        
        # 从C1单元格开始写入数据
        for idx, data in enumerate(unique_data):
            cell = worksheet.cell(row=1, column=3 + idx)  # C列是第3列
            cell.value = data
        
        # 保存更改
        book.save(excel_file_path)
        book.close()
        
        print(f"成功将 {len(unique_data)} 条数据写入到REGULATION工作表的C1单元格及后续位置")
    
    except Exception as e:
        print(f"发生错误：{str(e)}")

if __name__ == "__main__":
    write_unique_data_to_excel()