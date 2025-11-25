import pandas as pd
import json
import os
from openpyxl import load_workbook

def extract_regulation_first_row():
    """
    提取database.xlsx中REGULATION工作表第一行从C列到最后一列的数据，
    将空格替换为连字符，按首字母升序排序后再将数据重新写回Excel表
    """
    # 设置文件路径
    excel_file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'DataStorage', 'database.xlsx')
    json_output_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'DataStorage', 'regulation_first_row.json')
    
    # 检查Excel文件是否存在
    if not os.path.exists(excel_file_path):
        print(f"错误：Excel文件不存在于路径: {excel_file_path}")
        return
    
    try:
        # 读取Excel文件的REGULATION工作表
        # 不设置header，这样可以自己控制数据读取
        df = pd.read_excel(excel_file_path, sheet_name='REGULATION', header=None)
        
        # 检查是否有数据
        if df.empty:
            print("错误：工作表为空")
            return
        
        # 提取第一行（索引0）从C列（索引2）到最后一列的数据
        first_row_data = df.iloc[0, 2:].tolist()
        
        # 过滤掉空值并转换为字符串
        processed_data = []
        for value in first_row_data:
            if pd.notna(value):
                # 转换为字符串
                str_value = str(value)
                # 替换空格为连字符
                processed_value = str_value.replace(' ', '-')
                processed_data.append(processed_value)
        
        # 按首字母升序排序（首字母相同则按次字母升序）
        sorted_data = sorted(processed_data, key=lambda x: x.lower())  # 使用lower()确保不区分大小写
        
        # 保存原始数据为JSON文件
        original_data_clean = [value for value in first_row_data if pd.notna(value)]
        with open(json_output_path, 'w', encoding='utf-8') as json_file:
            json.dump(original_data_clean, json_file, ensure_ascii=False, indent=2)
        
        # 使用openpyxl将处理后的数据写回Excel文件
        wb = load_workbook(excel_file_path)
        ws = wb['REGULATION']
        
        # 从C1单元格开始写回排序后的数据
        for i, value in enumerate(sorted_data):
            col_index = i + 3  # C列对应索引3
            ws.cell(row=1, column=col_index).value = value
        
        # 保存修改后的Excel文件
        wb.save(excel_file_path)
        wb.close()
        
        print(f"成功提取{len(original_data_clean)}个数据项")
        print(f"原始数据已保存到: {json_output_path}")
        print(f"处理后的数据已按首字母升序排序并写回Excel文件")
        print(f"排序后的数据: {sorted_data}")
        
    except Exception as e:
        print(f"处理过程中出现错误: {str(e)}")

if __name__ == "__main__":
    extract_regulation_first_row()